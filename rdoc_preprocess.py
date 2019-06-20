
import os
from my_utils import Alphabet, pad_sequence
import csv
import codecs
from openpyxl.reader.excel import load_workbook
import nltk
nlp_tool = nltk.data.load('tokenizers/punkt/spanish.pickle')
from pytorch_pretrained_bert import BertTokenizer
from options import opt
from my_utils import my_tokenize
wp_tokenizer = BertTokenizer.from_pretrained(opt.bert_dir, do_lower_case=opt.do_lower_case)
import logging

class Document:
    def __init__(self):
        self.relevant_sentences = []
        self.abstract = ''
        self.pmid = ''
        self.title = ''
        self.category = ''
        self.sentences = []

class Sentence:
    def __init__(self):
        self.tokens = []
        self.start = -1
        self.end = -1
        self.text = ''
        self.label = ''

def load_data(data_dir, mode='train'):

    doc_num = 0
    sent_num = 0
    max_sent_length = 0
    min_sent_length = 9999
    total_sent_length = 0

    documents = []

    alphabet_category = Alphabet('category', True)
    print(os.listdir(data_dir))
    for input_file_name in os.listdir(data_dir):

        alphabet_category.add(input_file_name)

        wb = load_workbook(os.path.join(data_dir, input_file_name))

        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])

        for row_idx, row in enumerate(ws.rows):
            if row_idx == 0:
                continue # head
            document = Document()
            document.pmid = row[0].value
            document.title = row[1].value
            document.abstract = row[2].value
            if mode == 'train':
                document.relevant_sentences = parseReleventFromExcel(row[3].value)
            document.category = input_file_name

            all_sents_inds = []
            generator = nlp_tool.span_tokenize(document.abstract)
            for t in generator:
                all_sents_inds.append(t)

            for ind in range(len(all_sents_inds)):
                sentence = Sentence()
                sentence.start = all_sents_inds[ind][0]
                sentence.end = all_sents_inds[ind][1]

                offset = 0
                sentence.text = document.abstract[sentence.start:sentence.end]
                if len(document.relevant_sentences) != 0:
                    if sentence.text in document.relevant_sentences:
                        sentence.label = 'yes'
                    else:
                        sentence.label = 'no'
                else:
                    sentence.label = 'no'

                # replace due to nltk transfer " to other character, see https://github.com/nltk/nltk/issues/1630
                sentence.text = sentence.text.replace('"', " ")
                sentence.text = sentence.text.replace('\'', " ")
                for token_txt in my_tokenize(sentence.text):
                    token = {}
                    offset = sentence.text.find(token_txt, offset)
                    if offset == -1:
                        raise RuntimeError("can't find {} in '{}'".format(token_txt, sentence.text))

                    token['text'] = token_txt
                    token['start'] = sentence.start + offset
                    token['end'] = sentence.start + offset + len(token_txt)
                    token['wp'] = wp_tokenizer.tokenize(token_txt)
                    if len(token['wp']) == 0: # for some oov tokens (e.g., \x99), wp_tokenizer return a empty list
                        token['wp'] = ['[UNK]']

                    sentence.tokens.append(token)
                    offset += len(token_txt)

                document.sentences.append(sentence)
                sent_num += 1
                total_sent_length += len(sentence.tokens)
                if len(sentence.tokens) > max_sent_length:
                    max_sent_length = len(sentence.tokens)
                if len(sentence.tokens) < min_sent_length:
                    min_sent_length = len(sentence.tokens)

            documents.append(document)
            doc_num += 1


    logging.info("{} statistics".format(data_dir))
    logging.info("doc number {}, sent number {}".format(doc_num, sent_num))
    logging.info("avg sent length {}, max sent length {}, min sent length {}".format(total_sent_length//sent_num,
                                                                                     max_sent_length, min_sent_length))

    return documents, alphabet_category


def parseReleventFromExcel(txt):
    sentences = []

    offset = 0
    left_quote = txt.find("'", offset)
    offset = left_quote+1
    right_quote = txt.find("'", offset)
    offset = right_quote+1

    while left_quote != -1 and right_quote != -1:
        sentences.append(txt[left_quote + 1:right_quote])

        left_quote = txt.find("'", offset)
        offset = left_quote+1
        right_quote = txt.find("'", offset)
        offset = right_quote+1

    return sentences

def build_alphabet(documents, label_alphabet):
    for document in documents:
        for sentence in document.sentences:
            label_alphabet.add(sentence.label)


def prepare_instance(documents, alphabet_label, alphabet_category):
    instances = []

    for document in documents:

        instances_one_doc = prepare_instance_for_one_doc(document, alphabet_label, alphabet_category)

        instances.extend(instances_one_doc)

    return instances

# one sentence one instance
def prepare_instance_for_one_doc(document, alphabet_label, alphabet_category):
    instances = []

    for sentence in document.sentences:

        tokens = []

        for token in sentence.tokens:
            word_pieces = token['wp']
            tokens.extend(word_pieces)

        if len(tokens) > opt.len_max_seq-2: # for [CLS] and [SEP]
            tokens = tokens[:opt.len_max_seq-2]
            # labels = labels[:opt.len_max_seq-2]

        tokens.insert(0, '[CLS]')
        tokens.append('[SEP]')
        tokens = wp_tokenizer.convert_tokens_to_ids(tokens)

        instance = {'tokens': tokens, 'labels': alphabet_label.get_index(sentence.label),
                    'cate': alphabet_category.get_index(document.category)}

        instances.append(instance)

    return instances

import torch
def my_collate(x):
    tokens = [x_['tokens'] for x_ in x]
    labels = [x_['labels'] for x_ in x]
    cate = [x_['cate'] for x_ in x]
    lengths = [len(x_['tokens']) for x_ in x]
    max_len = max(lengths)
    mask = [[1] * length for length in lengths]
    sent_type = [[0]* length for length in lengths]

    tokens = pad_sequence(tokens, max_len)
    labels = torch.LongTensor(labels)
    cate = torch.LongTensor(cate)
    mask = pad_sequence(mask, max_len)
    sent_type = pad_sequence(sent_type, max_len)

    if opt.gpu >= 0 and torch.cuda.is_available():
        tokens = tokens.cuda(opt.gpu)
        labels = labels.cuda(opt.gpu)
        cate = cate.cuda(opt.gpu)
        mask = mask.cuda(opt.gpu)
        sent_type = sent_type.cuda(opt.gpu)

    return tokens, labels, mask, sent_type, cate

from torch.utils.data import DataLoader
from my_utils import MyDataset
def evaluate(documents, model, alphabet_label, alphabet_category, dump_dir):

    ct_predicted = 0
    ct_gold = 0
    ct_correct = 0

    all_pred_labels = []

    for document in documents:
        instances = prepare_instance_for_one_doc(document, alphabet_label, alphabet_category)

        data_loader = DataLoader(MyDataset(instances), opt.batch_size, shuffle=False, collate_fn=my_collate)

        pred_labels = []

        with torch.no_grad():
            model.eval()

            data_iter = iter(data_loader)
            num_iter = len(data_loader)
            sent_start = 0


            for i in range(num_iter):
                tokens, labels, mask, sent_type, cate = next(data_iter)

                logits = model.forward(cate, tokens, sent_type, mask)

                actual_batch_size = tokens.size(0)

                for batch_idx in range(actual_batch_size):

                    sent_logits = logits[batch_idx]

                    _, indices = torch.max(sent_logits, 0)

                    pred_labels.append(alphabet_label.get_instance(indices.item()))



                sent_start += actual_batch_size

        if len(document.relevant_sentences) != 0:
            p1, p2, p3 = count_tp(document.sentences, pred_labels)
        else:
            p1, p2, p3 = 0, 0, 0


        if dump_dir:
            all_pred_labels.append(pred_labels)

        ct_gold += p1
        ct_predicted += p2
        ct_correct += p3

    if ct_gold == 0 or ct_predicted == 0:
        precision = 0
        recall = 0
    else:
        precision = ct_correct * 1.0 / ct_predicted
        recall = ct_correct * 1.0 / ct_gold

    if precision+recall == 0:
        f_measure = 0
    else:
        f_measure = 2*precision*recall/(precision+recall)

    # dump results
    if dump_dir:
        dump_results(documents, all_pred_labels, dump_dir)

    return precision, recall, f_measure

# only consider 'yes'
def count_tp(gold, pred):

    ct_gold = 0
    ct_predict = 0
    ct_correct = 0

    for g in gold:
        if g.label=='yes':
            ct_gold += 1

    for i, p in enumerate(pred):
        if p=='yes':
            ct_predict += 1
            if p == gold[i].label:
                ct_correct += 1

    return ct_gold, ct_predict, ct_correct

from openpyxl import Workbook
def dump_results(documents, all_pred_labels, dump_dir):

    cate_documents = {}
    # the same category input into one directory
    for document_idx, document in enumerate(documents):
        pred_label = all_pred_labels[document_idx]

        if document.category in cate_documents:
            one_cate = cate_documents[document.category]
            one_cate.append({"document":document, "label":pred_label})
        else:
            one_cate = [{"document":document, "label":pred_label}]
            cate_documents[document.category] = one_cate

    # one category one xlsx
    for category, doc_labels in cate_documents.items():
        wb = Workbook()
        ws1 = wb.get_sheet_by_name('Sheet')

        ws1.cell(row=1, column=1, value='pmid')
        ws1.cell(row=1, column=2, value='title')
        ws1.cell(row=1, column=3, value='abstract')
        ws1.cell(row=1, column=4, value='Relevant Context')

        row_idx = 2
        for doc_label in doc_labels:
            document = doc_label['document']
            pred_label = doc_label['label']

            ws1.cell(row=row_idx, column=1, value=document.pmid)
            ws1.cell(row=row_idx, column=2, value=document.title)
            ws1.cell(row=row_idx, column=3, value=document.abstract)

            str_relevant = '{'
            all_no = True
            for sent_idx, sentence in enumerate(document.sentences):
                if pred_label[sent_idx] == 'yes':
                    str_relevant += "'"+sentence.text+"', "
                    all_no = False

            if all_no == False:
                str_relevant = str_relevant[:-2] # remove the last , and whitespace
            str_relevant += '}'
            ws1.cell(row=row_idx, column=4, value=str_relevant)

            row_idx += 1

        wb.save(os.path.join(dump_dir, category))


def translate_xlsx_to_txt(documents, path):
    category = ""
    with codecs.open(path, 'w', 'UTF-8') as fp:
        for document in documents:
            if category != document.category:
                fp.write(document.category[:document.category.find('.xlsx')]+"\n")
                category = document.category

            for sentence in document.relevant_sentences:
                fp.write(str(document.pmid)+"\t"+sentence+"\n")




